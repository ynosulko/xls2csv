#!/usr/bin/env python

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
from os import sys


def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    all_worksheets = workbook.sheetnames
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets


def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file, data_only=True)
    for worksheet_name in sheets:
        print("Exporting " + worksheet_name + " ...")

        try:
            worksheet = workbook[worksheet_name]
        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)

        your_csv_file = open(''.join([worksheet_name, '.csv']), 'wb')
        wr = csv.writer(your_csv_file, delimiter=';', quoting=csv.QUOTE_NONE)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wr.writerow(lrow)
        print(" ... done!")
    your_csv_file.close()


if not 2 <= len(sys.argv) <= 3:
    print("Call with " +
          sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
    sys.exit(1)
else:
    sheets = []
    if len(sys.argv) == 3:
        sheets = list(sys.argv[2].split(','))
    else:
        sheets = get_all_sheets(sys.argv[1])
    assert(sheets != None and len(sheets) > 0)
    csv_from_excel(sys.argv[1], sheets)
    
